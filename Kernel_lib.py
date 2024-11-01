#for this notebook to run, you'd need python 3, numpy and scipy
import numpy as np
import os
from scipy.linalg import sqrtm
from scipy.sparse import csr_matrix
from scipy.sparse.linalg import lsqr
from sklearn.model_selection import cross_validate
from scipy.sparse import csr_matrix, vstack, hstack, diags, save_npz
from sklearn.metrics import make_scorer
from sklearn.metrics import f1_score, precision_score, recall_score, roc_auc_score
import openpyxl
#import tensorflow as tf
#from numba import jit

np.random.seed(42)


def random_inv_matrix(matrixsize): 
    """
    Generate a strictly diagonally dominant random matrix of size matrixsize x matrixsize.

    :param matrixsize: the size of the matrix to generate
    :return: a strictly diagonally dominant random matrix of size matrixsize x matrixsize
    """
    R = np.random.rand(matrixsize, matrixsize)
    mx = np.sum(np.abs(R), axis=1)
    np.fill_diagonal(R, mx)
    return R

def random_dummy_features(n_samples):
    """
    Generate a matrix of dummy features with a random number of rows and a fixed number of columns.

    :param n_samples: the number of columns in the matrix
    :return: a matrix of dummy features with a random number of rows and n_samples columns
    """
    np.random.seed(42)
    M1_rand = int(np.random.randint(1, 21, 1)) #random int from 1-20
    dummy_features = np.zeros((M1_rand, n_samples))
    return dummy_features

def get_half_inv(R): 
    """
    Compute the half-inverse of a matrix.

    :param R: the matrix to compute the half-inverse of
    :return: the half-inverse R^{-1/2} of R
    """
    R_inv = np.linalg.inv(R)
    R_half_inv = sqrtm(R_inv)
    return R_half_inv

def randomize_matrix(input_matrix, R):
    """
    Randomize a matrix using the Nystroem method.

    :param input_matrix: the matrix to randomize
    :param R: the random matrix used in the Nystroem method
    :return: the randomized matrix
    """
    nystroem1 = input_matrix.transpose().dot(R)
    nystroem3 = get_half_inv(R.transpose().dot(R))
    return nystroem1.dot(nystroem3)


###############################################################################################
# LeftRight Approach
###############################################################################################

#  R^{1/2}
def get_R_half(R): 
    """
    R^{1/2}

    :param R: data matrix
    :return R_half: square root of data matrix
     """
    R_half = sqrtm(R)
    return R_half


def random_matrix(f, k):
    """
    full rank, invertible random matrix for some f < k

    :param f: number of features
    :param k: number of datapoints
    :return R: random matrix
     """
    np.random.seed(42)    
    while True: 
        R = np.random.rand(f, k)
        if np.linalg.matrix_rank(R) == min(f, k): #full rank and invertible
            return R
            break
            
#generate matrix of with left inverse masked data
def generate_data_prime(ds, N):
    """
    mask the data with random matrix and left inverse

    :param ds: matrix
    :param N: full rank random matrix
    :return ds_prime: masked data
     """
        
    N_linv = get_left_inverse(N)
    temp = get_R_half(N.dot(N.transpose())) 
    temp2 = N_linv.dot(temp) 
    ds_prime=ds.dot(temp2)
    ds_prime = ds_prime.real
    return ds_prime

def get_left_inverse(ds):
    """
    get left inverse

    :param ds: matrix
    :return N_linv: left inverse
     """
    
    N_linv = np.linalg.solve(ds.T.dot(ds), ds.T)
    return N_linv

def get_pseudo_inverse(ds):
    """
    get Moore–Penrose inverse (pseudo inverse) computed with SVD

    :param ds: matrix
    :return N_linv: pseudo inverse
     """
    # Perform SVD
    U, s, V = np.linalg.svd(ds)

    # Compute the left inverse of A
    A_inv = V @ np.linalg.inv(np.diag(s)) @ U.T

    return A_inv


def get_pseudo_inverse_pinv(ds):
    """
    get Moore–Penrose inverse (pseudo inverse) computed with pinv

    :param ds: matrix
    :return N_linv: pseudo inverse
     """
    A_inv = np.linalg.pinv(ds)
    return A_inv    


###############################################################################################
# Classification
###############################################################################################


def metrics_micro(clf, X, y, cv, seed=42):
    """
    Use k-fold cross-validation with cv folds to evaluate the classifier

    :param clf: sklearn classifier
    :param X: dataset
    :param y: labels
    :param cv: cross validation folds
    :param seed: random seed for reproducibility
    :print: mean and standard deviation of the F1 micro
    :print: mean and standard deviation of the ROC AUC scores
    :print: mean and standard deviation of the precision micro
    :print: mean and standard deviation of the recall micro
    :print: mean and standard deviation of the accuracy
     """
    
    # Set random seed for reproducibility
    np.random.seed(seed)

    # Use cross-validation to evaluate the classifier
    scoring = ['f1_micro', 'roc_auc_ovo', 'precision_micro', 'recall_micro', 'accuracy']
    scores = cross_validate(clf, X, y, cv=cv, scoring=scoring, return_train_score=True)

    # Print the mean and standard deviation of metrics
    print("F1 Score Micro: %0.2f (+/- %0.2f)" % (scores['test_f1_micro'].mean(), scores['test_f1_micro'].std() * 2))
    print("ROC AUC Score: %0.2f (+/- %0.2f)" % (scores['test_roc_auc_ovo'].mean(), scores['test_roc_auc_ovo'].std() * 2))
    print("Precision Micro Score: %0.2f (+/- %0.2f)" % (scores['test_precision_micro'].mean(), scores['test_precision_micro'].std() * 2))
    print("Recall Micro Score: %0.2f (+/- %0.2f)" % (scores['test_recall_micro'].mean(), scores['test_recall_micro'].std() * 2))
    print("Accuracy: %0.2f (+/- %0.2f)" % (scores['test_accuracy'].mean(), scores['test_accuracy'].std() * 2))

def metrics_macro(clf, X, y, cv):
    """
    Use k-fold cross-validation with cv folds to evaluate the classifier

    :param clf: sklearn classifier
    :param X: dataset
    :param y: labels
    :param cv: cross validation folds
    :print: mean and standard deviation of the F1 macro
    :print: mean and standard deviation of the ROC AUC scores
    :print: mean and standard deviation of the precision macro
    :print: mean and standard deviation of the recall macro
     """
    
    # Use cross-validation to evaluate the classifier
    scores = cross_validate(clf, X, y, cv = cv, scoring=['f1_macro', 'roc_auc_ovo', 'precision_macro', 'recall_macro',  'accuracy'], return_train_score=True)

    # Print the mean and standard deviation of metrics
    print("F1 Score macro: %0.2f (+/- %0.2f)" % (scores['test_f1_macro'].mean(), scores['test_f1_macro'].std() * 2))
    print("ROC AUC Score: %0.2f (+/- %0.2f)" % (scores['test_roc_auc_ovo'].mean(), scores['test_roc_auc_ovo'].std() * 2))
    print("Precision macro Score: %0.2f (+/- %0.2f)" % (scores['test_precision_macro'].mean(), scores['test_precision_macro'].std() * 2))
    print("Recall macro Score: %0.2f (+/- %0.2f)" % (scores['test_recall_macro'].mean(), scores['test_recall_macro'].std() * 2))
    print("Accuracy: %0.2f (+/- %0.2f)" % (scores['test_accuracy'].mean(), scores['test_accuracy'].std() * 2))

def dump_serverside_results(clf, X, y, cv, times, path):
    """
    Use k-fold cross-validation with cv folds to evaluate the classifier and write the results to the next available row in an Excel file

    :param clf: sklearn classifier
    :param X: dataset
    :param y: labels
    :param cv: cross validation folds
    :param times: list of communication_time: time taken to communicate the data,
        gram_time: time taken to calculate the Gram matrix,
        training_time: time taken to train the classifier
     """

    # Use cross-validation to evaluate the classifier
    scores = cross_validate(clf, X, y, cv=cv, scoring=['f1_micro', 'roc_auc_ovo', 'precision_micro', 'recall_micro'], return_train_score=True)

    # Create the Excel file if it does not exist
    if not os.path.exists(path):
        wb = openpyxl.Workbook()
        wb.save(path)

    # Open the Excel file
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Find the next available row
    next_row = ws.max_row + 1

    # Write the mean and standard deviation of each metric to the next available row
    ws.cell(row=next_row, column=1).value = scores['test_f1_micro'].mean()
    ws.cell(row=next_row, column=2).value = scores['test_roc_auc_ovo'].mean()
    ws.cell(row=next_row, column=3).value = scores['test_precision_micro'].mean()
    ws.cell(row=next_row, column=4).value = scores['test_recall_micro'].mean()
    ws.cell(row=next_row, column=5).value = scores['test_f1_micro'].std() * 2
    ws.cell(row=next_row, column=6).value = scores['test_roc_auc_ovo'].std() * 2
    ws.cell(row=next_row, column=7).value = scores['test_precision_micro'].std() * 2
    ws.cell(row=next_row, column=8).value = scores['test_recall_micro'].std() * 2
    ws.cell(row=next_row, column=9).value = times[0]
    ws.cell(row=next_row, column=10).value = times[1]
    ws.cell(row=next_row, column=11).value = times[2]

    # Save the Excel file
    wb.save(path)

def dump_clientside_results(masking_time, path):
    """
    Dumps the time to mask the data into the result escel sheet 

    :param times: masking time
     """

    # Create the Excel file if it does not exist
    if not os.path.exists(path):
        wb = openpyxl.Workbook()
        wb.save(path)

    # Open the Excel file
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Find the next available row
    next_row = ws.max_row + 1

    # Write the mean and standard deviation of each metric to the next available row
    ws.cell(row=next_row, column=1).value = masking_time
    # Save the Excel file
    wb.save(path)

# Polynomial Kernel fpr Gram Matrix
def polynomial_kernel(dataset, degree):
    """
    Compute kernel value of each entry in Gram Matrix

    :param dataset: dataset of size features*datapoints
    :param degree: degree for polynom kernel
    :return K: kerneled dataset 
     """
    
    m1,m2 = dataset.shape
    K = np.zeros((m1, m2))
    for i in range(m1):
        for j in range(m2):
            K[i,j] = (dataset[i,j] + 1)**degree 
    return K


def laplace_kernel(dataset, gamma):
    """
    Compute Laplace (RBF) kernel value of each entry in Gram Matrix

    :param dataset: dataset of size features*datapoints
    :param gamma: gamma parameter for Laplace kernel
    :return K: kerneled dataset 
    """
    
    m1, m2 = dataset.shape
    K = np.zeros((m1, m2))
    for i in range(m1):
        for j in range(m2):
            diff = dataset[i, :] - dataset[j, :]
            K[i, j] = np.exp(-gamma * np.linalg.norm(diff))
    return K

###############################################################################################
# Simulating multiple nodes
###############################################################################################

def partition_dataset(dataset, labels, n):
    """
    Computes Gram Matrix

    :param dataset: dataset of size features*datapoints
    :param labels: list of labels
    :param n: count of partitions 
    :return partitions: list of n arrays 
    :return labels: array of labels, cropped depending on excess
     """
        
  # Convert the dataset and labels to numpy arrays, if they aren't already
    dataset = np.asarray(dataset)
    labels = np.asarray(labels)
  
  # Calculate the size of each partition
    partition_size = len(dataset) // n

  # If the dataset size is not evenly divisible by n, discard the excess elements
    excess = len(dataset) % n
    if excess > 0:
        dataset = dataset[:-excess]
        labels = labels[:-excess]

  # Use np.split to partition the dataset and labels into n equal parts
    partitions = np.split(dataset, n)
  
    return partitions, labels


def compute_gram_matrix(*dataset):
    """
    Computes Gram Matrix

    :param dataset: list of m arrays of size features*datapoints, numbers of features have to be the same
    :return: Gram Matrix, Size n*m
    """

    if isinstance(dataset[0], list):
        # The first argument is a list, so treat all of the arguments as individual data arrays
        data_arrays = dataset[0]
    else:
        # The first argument is not a list, so treat all of the arguments as individual data arrays
        data_arrays = dataset
    
    rows = sum(a.shape[0] for a in data_arrays)
    cols = rows
    full = np.zeros((rows, cols))

    row_offset = 0
    for i, A in enumerate(data_arrays):
        col_offset = 0
        for j, B in enumerate(data_arrays):
            full[row_offset:row_offset+A.shape[0], col_offset:col_offset+B.shape[0]] = np.dot(A, B.T)
            col_offset += B.shape[0]
        row_offset += A.shape[0]

    return full
   
def update_gram_matrix(full, *dataset, new_dataset):
    """
    Extends Gram Matrix with new dot products

    :param full: existing Gram matrix, size n*m
    :param datasets: m arrays of size features*datapoints, numbers of features have to be the same
    :param new_dataset: new array of size features*datapoints, number of features has to be the same as in *datasets
    :return: extended Gram Matrix, Size n*(m+1)
    """

    data_arrays = list(dataset) + [new_dataset]

    rows = full.shape[0]
    cols = rows + new_dataset.shape[0]
    extended = np.zeros((rows, cols))
    extended[:, :-new_dataset.shape[0]] = full

    row_offset = 0
    for i, A in enumerate(data_arrays):
        col_offset = full.shape[1]
        B = new_dataset
        extended[row_offset:row_offset+A.shape[0], col_offset:col_offset+B.shape[0]] = np.dot(A, B.T)
        row_offset += A.shape[0]

    return extended
  

###############################################################################################
# Other Methods
###############################################################################################


def perform_rlt(data):
    # Generate a random linear transformation matrix
    transformation_matrix = np.random.rand(data.shape[1], data.shape[1])

    # Apply the random linear transformation to the data
    transformed_data = np.dot(data, transformation_matrix)

    return transformed_data


def perform_random_kernel(data):
    # Generate random kernel values
    random_kernel_values = np.random.rand(data.shape[1], data.shape[0])

    # Apply the random kernel to the data
    transformed_data = np.dot(data, random_kernel_values)

    return transformed_data


